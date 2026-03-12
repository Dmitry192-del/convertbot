"""
Telegram File Converter Bot

Фичи:
  - Конвертация файлов (PDF/DOCX/XLSX/PPTX/JPG/PNG и др.)
  - Переименование файлов
  - Сжатие PDF
  - Статистика /stats, история /history
  - Админка /admin
  - Прогресс-бар при конвертации

Setup:
    pip install python-telegram-bot python-dotenv pdf2docx docx2pdf Pillow
                img2pdf pdfplumber openpyxl reportlab pypdf

.env:
    BOT_TOKEN=ваш_токен
    ADMIN_ID=ваш_telegram_id        # узнать у @userinfobot
"""

import os
from dotenv import load_dotenv
load_dotenv()

import uuid
import json
import time
import asyncio
import logging
import tempfile
import subprocess
from pathlib import Path
from datetime import datetime

from telegram import Update, InlineKeyboardButton, InlineKeyboardMarkup, BotCommand
from telegram.ext import (
    Application,
    CommandHandler,
    MessageHandler,
    CallbackQueryHandler,
    ContextTypes,
    ConversationHandler,
    filters,
)

# ── Logging ────────────────────────────────────────────────────────────────────
logging.basicConfig(
    format="%(asctime)s | %(levelname)s | %(name)s | %(message)s",
    level=logging.INFO,
)
logger = logging.getLogger(__name__)

# ── Config ─────────────────────────────────────────────────────────────────────
ADMIN_ID    = int(os.environ.get("ADMIN_ID", "0"))
STATS_FILE  = Path("stats.json")
MAX_FILE_MB = 20
SESSION_TTL = 60 * 60  # сессии живут 1 час (в секундах)

# ── ConversationHandler states ─────────────────────────────────────────────────
WAITING_NEW_NAME = 1

# ── Conversion map ─────────────────────────────────────────────────────────────
CONVERSIONS: dict[tuple[str, str], str] = {
    ("jpg",  "pdf"):  "JPG → PDF",
    ("jpeg", "pdf"):  "JPEG → PDF",
    ("png",  "pdf"):  "PNG → PDF",
    ("docx", "pdf"):  "WORD → PDF",
    ("doc",  "pdf"):  "DOC → PDF",
    ("pptx", "pdf"):  "POWERPOINT → PDF",
    ("ppt",  "pdf"):  "PPT → PDF",
    ("xlsx", "pdf"):  "EXCEL → PDF",
    ("xls",  "pdf"):  "XLS → PDF",
    ("html", "pdf"):  "HTML → PDF",
    ("pdf",  "docx"): "PDF → WORD",
    ("pdf",  "xlsx"): "PDF → EXCEL",
}

SUPPORTED_INPUTS: dict[str, list[str]] = {}
for (src, tgt) in CONVERSIONS:
    SUPPORTED_INPUTS.setdefault(src, []).append(tgt)

FORMAT_EMOJI = {
    "pdf": "📕", "docx": "📘", "doc": "📘",
    "pptx": "📙", "ppt": "📙",
    "xlsx": "📗", "xls": "📗",
    "jpg": "🖼", "jpeg": "🖼", "png": "🖼",
    "html": "🌐",
}

def fmt_emoji(ext: str) -> str:
    return FORMAT_EMOJI.get(ext, "📄")


# ── Statistics ─────────────────────────────────────────────────────────────────

def _load_stats() -> dict:
    if STATS_FILE.exists():
        try:
            return json.loads(STATS_FILE.read_text(encoding="utf-8"))
        except Exception:
            pass
    return {"users": {}, "total_conversions": 0, "total_renames": 0, "total_compressions": 0}


def _save_stats(data: dict) -> None:
    STATS_FILE.write_text(json.dumps(data, ensure_ascii=False, indent=2), encoding="utf-8")


def record_action(user_id: int, username: str, action: str, detail: str = "") -> None:
    """action: 'convert' | 'rename' | 'compress'"""
    data = _load_stats()
    uid  = str(user_id)

    if uid not in data["users"]:
        data["users"][uid] = {
            "username": username,
            "first_seen": datetime.now().isoformat(timespec="seconds"),
            "conversions": 0,
            "renames": 0,
            "compressions": 0,
            "history": [],
        }

    u = data["users"][uid]
    u["username"] = username  # обновляем на случай смены ника

    now = datetime.now().strftime("%d.%m.%Y %H:%M")
    if action == "convert":
        u["conversions"] += 1
        data["total_conversions"] += 1
        u["history"].insert(0, f"{now} 🔄 {detail}")
    elif action == "rename":
        u["renames"] += 1
        data["total_renames"] += 1
        u["history"].insert(0, f"{now} ✏️ {detail}")
    elif action == "compress":
        u["compressions"] += 1
        data["total_compressions"] += 1
        u["history"].insert(0, f"{now} 🗜 {detail}")

    u["history"] = u["history"][:20]  # храним последние 20
    _save_stats(data)


# ── Progress bar helper ────────────────────────────────────────────────────────

async def animated_progress(
    context: ContextTypes.DEFAULT_TYPE,
    chat_id: int,
    message_id: int,
    label: str,
    stop_event: asyncio.Event,
) -> None:
    """Крутит прогресс-бар пока stop_event не установлен."""
    frames = [
        "▱▱▱▱▱▱▱▱",
        "▰▱▱▱▱▱▱▱",
        "▰▰▱▱▱▱▱▱",
        "▰▰▰▱▱▱▱▱",
        "▰▰▰▰▱▱▱▱",
        "▰▰▰▰▰▱▱▱",
        "▰▰▰▰▰▰▱▱",
        "▰▰▰▰▰▰▰▱",
        "▰▰▰▰▰▰▰▰",
    ]
    i = 0
    while not stop_event.is_set():
        frame = frames[i % len(frames)]
        try:
            await context.bot.edit_message_text(
                chat_id=chat_id,
                message_id=message_id,
                text=f"⏳ *{label}*\n\n`{frame}`",
                parse_mode="Markdown",
            )
        except Exception:
            pass
        i += 1
        await asyncio.sleep(2.0)


# ── LibreOffice helper ─────────────────────────────────────────────────────────

def libreoffice_convert(src: Path, tgt_ext: str, dst: Path) -> None:
    result = subprocess.run(
        ["libreoffice", "--headless", "--convert-to", tgt_ext,
         "--outdir", str(dst.parent), str(src)],
        capture_output=True, text=True, timeout=120,
    )
    if result.returncode != 0:
        raise RuntimeError(f"LibreOffice: {result.stderr.strip()}")
    generated = dst.parent / (src.stem + f".{tgt_ext}")
    if generated.exists() and generated != dst:
        generated.rename(dst)


# ── Conversion functions ───────────────────────────────────────────────────────

def conv_image_to_pdf(src: Path, dst: Path) -> None:
    try:
        import img2pdf
        with open(dst, "wb") as f:
            f.write(img2pdf.convert(str(src)))
    except Exception:
        from PIL import Image
        Image.open(src).convert("RGB").save(str(dst), "PDF", resolution=150)


def _register_cyrillic_font():
    from reportlab.pdfbase import pdfmetrics
    from reportlab.pdfbase.ttfonts import TTFont
    import glob

    candidates = [
        # Linux (Railway/Render/Ubuntu)
        "/usr/share/fonts/truetype/dejavu/DejaVuSans.ttf",
        "/usr/share/fonts/truetype/liberation/LiberationSans-Regular.ttf",
        "/usr/share/fonts/truetype/freefont/FreeSans.ttf",
        "/usr/share/fonts/truetype/ubuntu/Ubuntu-R.ttf",
        # Windows
        "C:/Windows/Fonts/arial.ttf",
        "C:/Windows/Fonts/calibri.ttf",
        "C:/Windows/Fonts/times.ttf",
        # macOS
        "/System/Library/Fonts/Supplemental/Arial.ttf",
        "/Library/Fonts/Arial.ttf",
    ]
    candidates += glob.glob("C:/Windows/Fonts/*.ttf")
    candidates += glob.glob("/usr/share/fonts/truetype/**/*.ttf", recursive=True)

    for path in candidates:
        try:
            pdfmetrics.registerFont(TTFont("CyrFont", path))
            pdfmetrics.registerFont(TTFont("CyrFont-Bold", path))
            return "CyrFont", "CyrFont-Bold"
        except Exception:
            continue
    return "Helvetica", "Helvetica-Bold"


def conv_xlsx_to_pdf(src: Path, dst: Path) -> None:
    import openpyxl
    from reportlab.lib.pagesizes import A4, landscape
    from reportlab.lib import colors
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib.units import mm

    font_name, font_bold = _register_cyrillic_font()
    wb = openpyxl.load_workbook(src, data_only=True)
    base_styles = getSampleStyleSheet()
    heading_style = ParagraphStyle("CyrHeading", parent=base_styles["Heading2"],
                                   fontName=font_bold, fontSize=11)
    cell_style = ParagraphStyle("CyrCell", parent=base_styles["Normal"],
                                fontName=font_name, fontSize=8)

    doc = SimpleDocTemplate(str(dst), pagesize=landscape(A4),
                            leftMargin=10*mm, rightMargin=10*mm,
                            topMargin=10*mm, bottomMargin=10*mm)
    elements = []

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        elements.append(Paragraph(sheet_name, heading_style))
        elements.append(Spacer(1, 3*mm))
        rows = list(ws.iter_rows(values_only=True))
        if not rows:
            continue
        table_data = [
            [Paragraph(str(cell) if cell is not None else "", cell_style) for cell in row]
            for row in rows
        ]
        col_count = max(len(r) for r in table_data)
        col_width = (landscape(A4)[0] - 20*mm) / max(col_count, 1)
        t = Table(table_data, colWidths=[col_width] * col_count, repeatRows=1)
        t.setStyle(TableStyle([
            ("BACKGROUND",     (0, 0), (-1, 0),  colors.HexColor("#4472C4")),
            ("TEXTCOLOR",      (0, 0), (-1, 0),  colors.white),
            ("FONTNAME",       (0, 0), (-1, 0),  font_bold),
            ("FONTSIZE",       (0, 0), (-1, -1), 8),
            ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#EEF2FF")]),
            ("GRID",           (0, 0), (-1, -1), 0.3, colors.HexColor("#CCCCCC")),
            ("ALIGN",          (0, 0), (-1, -1), "LEFT"),
            ("VALIGN",         (0, 0), (-1, -1), "MIDDLE"),
            ("TOPPADDING",     (0, 0), (-1, -1), 3),
            ("BOTTOMPADDING",  (0, 0), (-1, -1), 3),
        ]))
        elements.append(t)
        elements.append(Spacer(1, 6*mm))

    doc.build(elements)


def conv_docx_to_pdf_pure(src: Path, dst: Path) -> None:
    """DOCX → PDF через docx2pdf (Windows) или docx2python + reportlab (Linux)."""
    # Сначала пробуем docx2pdf (работает на Windows с MS Office)
    try:
        from docx2pdf import convert
        convert(str(src), str(dst))
        return
    except Exception:
        pass

    # Fallback: извлекаем текст через python-docx и рендерим через reportlab
    try:
        import docx
        from reportlab.lib.pagesizes import A4
        from reportlab.lib.units import mm
        from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer
        from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle

        font_name, font_bold = _register_cyrillic_font()
        base_styles = getSampleStyleSheet()
        normal = ParagraphStyle("N", parent=base_styles["Normal"], fontName=font_name, fontSize=11, leading=16)
        h1     = ParagraphStyle("H1", parent=base_styles["Heading1"], fontName=font_bold, fontSize=14, leading=20)

        document = docx.Document(str(src))
        doc = SimpleDocTemplate(str(dst), pagesize=A4,
                                leftMargin=20*mm, rightMargin=20*mm,
                                topMargin=20*mm, bottomMargin=20*mm)
        elements = []
        for para in document.paragraphs:
            text = para.text.strip()
            if not text:
                elements.append(Spacer(1, 4*mm))
                continue
            style = h1 if para.style.name.startswith("Heading") else normal
            elements.append(Paragraph(text, style))
        doc.build(elements)
        return
    except Exception as e:
        raise RuntimeError(f"Не удалось конвертировать DOCX: {e}")


def conv_html_to_pdf_pure(src: Path, dst: Path) -> None:
    """HTML → PDF: извлекаем текст и рендерим через reportlab."""
    import re
    from reportlab.lib.pagesizes import A4
    from reportlab.lib.units import mm
    from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle

    font_name, _ = _register_cyrillic_font()
    base_styles = getSampleStyleSheet()
    normal = ParagraphStyle("HN", parent=base_styles["Normal"],
                            fontName=font_name, fontSize=11, leading=16)

    with open(src, "r", encoding="utf-8", errors="ignore") as f:
        html = f.read()

    # Убираем теги, оставляем текст
    text = re.sub(r"<[^>]+>", " ", html)
    text = re.sub(r"&nbsp;", " ", text)
    text = re.sub(r"&amp;", "&", text)
    text = re.sub(r"&lt;", "<", text)
    text = re.sub(r"&gt;", ">", text)
    lines = [l.strip() for l in text.splitlines() if l.strip()]

    doc = SimpleDocTemplate(str(dst), pagesize=A4,
                            leftMargin=20*mm, rightMargin=20*mm,
                            topMargin=20*mm, bottomMargin=20*mm)
    elements = []
    for line in lines:
        elements.append(Paragraph(line, normal))
        elements.append(Spacer(1, 2*mm))
    doc.build(elements)


def conv_office_to_pdf(src: Path, dst: Path) -> None:
    ext = src.suffix.lower()
    if ext in (".xlsx", ".xls"):
        conv_xlsx_to_pdf(src, dst)
        return
    if ext in (".docx", ".doc"):
        conv_docx_to_pdf_pure(src, dst)
        return
    if ext == ".html":
        conv_html_to_pdf_pure(src, dst)
        return
    # PPTX/PPT — пробуем LibreOffice, если нет — сообщаем
    try:
        libreoffice_convert(src, "pdf", dst)
    except FileNotFoundError:
        raise RuntimeError("Конвертация PPTX→PDF требует LibreOffice. На текущем сервере он недоступен.")


def conv_pdf_to_docx(src: Path, dst: Path) -> None:
    from pdf2docx import Converter
    cv = Converter(str(src))
    cv.convert(str(dst), start=0, end=None)
    cv.close()


def conv_pdf_to_xlsx(src: Path, dst: Path) -> None:
    try:
        import pdfplumber, openpyxl
    except ImportError:
        raise RuntimeError("pip install pdfplumber openpyxl")

    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    with pdfplumber.open(str(src)) as pdf:
        for i, page in enumerate(pdf.pages, 1):
            ws = wb.create_sheet(title=f"Стр {i}")
            tables = page.extract_tables()
            if tables:
                row_idx = 1
                for table in tables:
                    for row in table:
                        for col_idx, cell in enumerate(row, 1):
                            ws.cell(row=row_idx, column=col_idx, value=cell or "")
                        row_idx += 1
                    row_idx += 1
            else:
                text = page.extract_text() or ""
                for row_idx, line in enumerate(text.splitlines(), 1):
                    ws.cell(row=row_idx, column=1, value=line)

    wb.save(str(dst))


def conv_compress_pdf(src: Path, dst: Path) -> tuple[int, int]:
    """Сжимает PDF через PyMuPDF (fitz) — реальное сжатие изображений и потоков.
    Возвращает (размер до, размер после) в байтах."""
    try:
        import fitz  # PyMuPDF
        doc = fitz.open(str(src))
        # Сжимаем каждое изображение внутри PDF
        for page in doc:
            for img in page.get_images(full=True):
                xref = img[0]
                try:
                    pix = fitz.Pixmap(doc, xref)
                    if pix.n >= 5:          # CMYK → RGB
                        pix = fitz.Pixmap(fitz.csRGB, pix)
                    compressed = fitz.Pixmap(pix)
                    doc.update_stream(xref, compressed.tobytes("jpeg", jpg_quality=60))
                except Exception:
                    pass
        # Сохраняем с дефлейт-компрессией и очисткой мусора
        doc.save(
            str(dst),
            garbage=4,          # максимальная очистка мусора
            deflate=True,       # сжатие потоков
            deflate_images=True,
            deflate_fonts=True,
            clean=True,
        )
        doc.close()
    except ImportError:
        # Fallback на pypdf если PyMuPDF не установлен
        from pypdf import PdfWriter, PdfReader
        reader = PdfReader(str(src))
        writer = PdfWriter()
        for page in reader.pages:
            page.compress_content_streams()
            writer.add_page(page)
        with open(dst, "wb") as f:
            writer.write(f)

    return src.stat().st_size, dst.stat().st_size


def conv_compress_pdf_quality(src: Path, dst: Path, jpg_quality: int = 60) -> tuple[int, int]:
    """Обёртка над conv_compress_pdf с настраиваемым качеством JPEG."""
    try:
        import fitz
        doc = fitz.open(str(src))
        for page in doc:
            for img in page.get_images(full=True):
                xref = img[0]
                try:
                    pix = fitz.Pixmap(doc, xref)
                    if pix.n >= 5:
                        pix = fitz.Pixmap(fitz.csRGB, pix)
                    doc.update_stream(xref, pix.tobytes("jpeg", jpg_quality=jpg_quality))
                except Exception:
                    pass
        doc.save(str(dst), garbage=4, deflate=True,
                 deflate_images=True, deflate_fonts=True, clean=True)
        doc.close()
    except ImportError:
        from pypdf import PdfWriter, PdfReader
        reader = PdfReader(str(src))
        writer = PdfWriter()
        for page in reader.pages:
            page.compress_content_streams()
            writer.add_page(page)
        with open(dst, "wb") as f:
            writer.write(f)
    return src.stat().st_size, dst.stat().st_size


def do_convert(src: Path, src_ext: str, tgt_ext: str, dst: Path) -> Path:
    key = (src_ext, tgt_ext)
    image_exts  = {"jpg", "jpeg", "png"}
    office_exts = {"docx", "doc", "pptx", "ppt", "xlsx", "xls", "html"}

    if src_ext in image_exts and tgt_ext == "pdf":
        conv_image_to_pdf(src, dst)
    elif src_ext in office_exts and tgt_ext == "pdf":
        conv_office_to_pdf(src, dst)
    elif key == ("pdf", "docx"):
        conv_pdf_to_docx(src, dst)
    elif key == ("pdf", "xlsx"):
        conv_pdf_to_xlsx(src, dst)
    else:
        raise ValueError(f"Неизвестная конвертация: {src_ext} → {tgt_ext}")

    return dst


# ── Keyboards ──────────────────────────────────────────────────────────────────

def action_keyboard(key: str, ext: str) -> InlineKeyboardMarkup:
    has_convert  = ext in SUPPORTED_INPUTS
    has_compress = ext == "pdf"
    buttons = []
    row = []
    if has_convert:
        row.append(InlineKeyboardButton("🔄 Конвертировать", callback_data=f"act|convert|{key}"))
    row.append(InlineKeyboardButton("✏️ Переименовать", callback_data=f"act|rename|{key}"))
    buttons.append(row)
    if has_compress:
        buttons.append([InlineKeyboardButton("🗜 Сжать PDF", callback_data=f"act|compress|{key}")])
    return InlineKeyboardMarkup(buttons)


def convert_keyboard(key: str, ext: str) -> InlineKeyboardMarkup:
    targets = SUPPORTED_INPUTS.get(ext, [])
    buttons = [
        [InlineKeyboardButton(
            f"{fmt_emoji(tgt)}  {CONVERSIONS[(ext, tgt)]}",
            callback_data=f"conv|{key}|{tgt}",
        )]
        for tgt in targets
    ]
    buttons.append([InlineKeyboardButton("⬅️ Назад", callback_data=f"act|back|{key}")])
    return InlineKeyboardMarkup(buttons)


# ── Commands ───────────────────────────────────────────────────────────────────

async def cmd_start(update: Update, context: ContextTypes.DEFAULT_TYPE) -> None:
    lines = "\n".join(
        f"  {fmt_emoji(s)} {label}" for (s, _), label in CONVERSIONS.items()
    )
    await update.message.reply_text(
        "👋 *Привет! Я бот-конвертер файлов.*\n\n"
        "Отправь мне файл и выбери действие:\n"
        "🔄 *Конвертировать* — сменить формат\n"
        "✏️ *Переименовать* — задать новое имя\n"
        "🗜 *Сжать PDF* — уменьшить размер файла\n\n"
        f"*Поддерживаемые конвертации:*\n{lines}",
        parse_mode="Markdown",
    )


async def cmd_help(update: Update, context: ContextTypes.DEFAULT_TYPE) -> None:
    await cmd_start(update, context)


async def cmd_stats(update: Update, context: ContextTypes.DEFAULT_TYPE) -> None:
    uid  = str(update.effective_user.id)
    data = _load_stats()
    u    = data["users"].get(uid)

    if not u:
        await update.message.reply_text("📊 Ты ещё не использовал бота. Отправь файл!")
        return

    await update.message.reply_text(
        f"📊 *Твоя статистика*\n\n"
        f"🔄 Конвертаций: *{u['conversions']}*\n"
        f"✏️ Переименований: *{u['renames']}*\n"
        f"🗜 Сжатий PDF: *{u['compressions']}*\n"
        f"📅 Первый запуск: {u['first_seen']}",
        parse_mode="Markdown",
    )


async def cmd_history(update: Update, context: ContextTypes.DEFAULT_TYPE) -> None:
    uid  = str(update.effective_user.id)
    data = _load_stats()
    u    = data["users"].get(uid)

    if not u or not u.get("history"):
        await update.message.reply_text("📋 История пуста.")
        return

    lines = "\n".join(f"`{h}`" for h in u["history"][:10])
    await update.message.reply_text(
        f"📋 *Последние 10 операций:*\n\n{lines}",
        parse_mode="Markdown",
    )


async def cmd_admin(update: Update, context: ContextTypes.DEFAULT_TYPE) -> None:
    user_id = update.effective_user.id
    logger.info(f"[ADMIN] user_id={user_id}, ADMIN_ID={ADMIN_ID}, match={user_id == ADMIN_ID}")
    if user_id != ADMIN_ID:
        await update.message.reply_text(f"⛔ Нет доступа. Твой ID: {user_id}")
        return

    data  = _load_stats()
    users = data["users"]
    total = len(users)

    top = sorted(users.items(),
                 key=lambda x: x[1]["conversions"] + x[1]["renames"] + x[1]["compressions"],
                 reverse=True)[:5]

    top_lines = "\n".join(
        f"  {i+1}. @{v['username'] or uid} — "
        f"{v['conversions']}🔄 {v['renames']}✏️ {v['compressions']}🗜"
        for i, (uid, v) in enumerate(top)
    ) or "  —"

    # Экранируем username чтобы не сломать Markdown
    safe_top = "\n".join(
        f"  {i+1}. @{v['username'].replace('_', chr(95)) or uid} — "
        f"{v['conversions']}🔄 {v['renames']}✏️ {v['compressions']}🗜"
        for i, (uid, v) in enumerate(top)
    ) or "  —"

    await update.message.reply_text(
        f"🛠 Админ-панель\n\n"
        f"👥 Пользователей: {total}\n"
        f"🔄 Конвертаций всего: {data['total_conversions']}\n"
        f"✏️ Переименований всего: {data['total_renames']}\n"
        f"🗜 Сжатий всего: {data['total_compressions']}\n\n"
        f"Топ-5 активных:\n{safe_top}",
    )


async def cmd_cancel(update: Update, context: ContextTypes.DEFAULT_TYPE) -> int:
    context.user_data.clear()
    await update.message.reply_text("❌ Отменено.")
    return ConversationHandler.END


# ── File handlers ──────────────────────────────────────────────────────────────

async def handle_document(update: Update, context: ContextTypes.DEFAULT_TYPE) -> None:
    doc       = update.message.document
    file_name = doc.file_name or "file"
    ext       = Path(file_name).suffix.lstrip(".").lower()
    size_mb   = (doc.file_size or 0) / 1024 / 1024

    if size_mb > MAX_FILE_MB:
        await update.message.reply_text(
            f"❌ Файл слишком большой ({size_mb:.1f} МБ).\n"
            f"Максимум: {MAX_FILE_MB} МБ."
        )
        return

    key = uuid.uuid4().hex[:12]
    context.bot_data[key] = {"file_id": doc.file_id, "ext": ext, "name": file_name}

    await update.message.reply_text(
        f"{fmt_emoji(ext)} Получил *{file_name}* ({size_mb:.1f} МБ)\n\nЧто хочешь сделать?",
        reply_markup=action_keyboard(key, ext),
        parse_mode="Markdown",
    )


async def handle_photo(update: Update, context: ContextTypes.DEFAULT_TYPE) -> None:
    photo = update.message.photo[-1]
    key   = uuid.uuid4().hex[:12]
    context.bot_data[key] = {"file_id": photo.file_id, "ext": "jpg", "name": "photo.jpg"}
    await update.message.reply_text(
        "🖼 Получил фото. Что хочешь сделать?",
        reply_markup=action_keyboard(key, "jpg"),
    )


# ── Callback handler ───────────────────────────────────────────────────────────

async def handle_callback(update: Update, context: ContextTypes.DEFAULT_TYPE):
    query = update.callback_query
    await query.answer()
    data  = query.data
    user  = update.effective_user

    # ── act|action|key ──
    if data.startswith("act|"):
        _, action, key = data.split("|", 2)
        info = context.bot_data.get(key)
        if not info:
            await query.edit_message_text("❌ Сессия устарела. Отправь файл заново.")
            return

        if action == "convert":
            ext = info["ext"]
            if ext not in SUPPORTED_INPUTS:
                await query.edit_message_text(
                    f"😕 Конвертация для *{ext.upper()}* не поддерживается.",
                    parse_mode="Markdown",
                )
                return
            await query.edit_message_text(
                f"{fmt_emoji(ext)} *{info['name']}*\n\nВыбери формат конвертации:",
                reply_markup=convert_keyboard(key, ext),
                parse_mode="Markdown",
            )

        elif action == "rename":
            context.user_data["rename_key"] = key
            await query.edit_message_text(
                f"✏️ Введи новое имя для *{info['name']}*\n"
                f"_(расширение можно не писать)_\n\n/cancel — отмена",
                parse_mode="Markdown",
            )
            return WAITING_NEW_NAME

        elif action == "compress":
            if info["ext"] != "pdf":
                await query.edit_message_text("❌ Сжатие доступно только для PDF.")
                return

            # Показываем меню сжатия с кнопкой Назад
            await query.edit_message_text(
                f"🗜 *Сжать PDF*\n\n📄 Файл: *{info['name']}*\n\nВыбери уровень сжатия:",
                reply_markup=InlineKeyboardMarkup([
                    [InlineKeyboardButton("🟢 Лёгкое (лучше качество)", callback_data=f"cmp|light|{key}")],
                    [InlineKeyboardButton("🟡 Среднее (баланс)",        callback_data=f"cmp|medium|{key}")],
                    [InlineKeyboardButton("🔴 Сильное (меньше размер)", callback_data=f"cmp|heavy|{key}")],
                    [InlineKeyboardButton("⬅️ Назад",                   callback_data=f"act|back|{key}")],
                ]),
                parse_mode="Markdown",
            )
            return

        elif action == "compress_do":
            compress_quality = context.user_data.pop("compress_quality", 75)
            # Запускаем прогресс-бар
            stop_event = asyncio.Event()
            progress_task = asyncio.create_task(
                animated_progress(context, query.message.chat_id,
                                  query.message.message_id, "Сжимаю PDF", stop_event)
            )

            try:
                with tempfile.TemporaryDirectory() as tmpdir:
                    tmp      = Path(tmpdir)
                    src_path = tmp / "input.pdf"
                    dst_path = tmp / "compressed.pdf"

                    tg_file = await context.bot.get_file(info["file_id"])
                    await tg_file.download_to_drive(str(src_path))

                    size_before, size_after = await asyncio.get_event_loop().run_in_executor(
                        None, conv_compress_pdf_quality, src_path, dst_path, compress_quality
                    )

                    stop_event.set()
                    await progress_task

                    saved    = size_before - size_after
                    saved_kb = saved / 1024
                    ratio    = (1 - size_after / size_before) * 100 if size_before else 0

                    stem     = Path(info["name"]).stem
                    out_name = f"{stem}_compressed.pdf"

                    await query.edit_message_text(
                        f"✅ Сжато!\n\n"
                        f"📦 До: *{size_before/1024:.0f} КБ*\n"
                        f"📦 После: *{size_after/1024:.0f} КБ*\n"
                        f"💾 Сэкономлено: *{saved_kb:.0f} КБ ({ratio:.0f}%)*",
                        parse_mode="Markdown",
                    )

                    with open(dst_path, "rb") as f:
                        await context.bot.send_document(
                            chat_id=query.message.chat_id,
                            document=f,
                            filename=out_name,
                            caption=f"🗜 {info['name']} → {out_name}",
                        )

                    record_action(user.id, user.username or "",
                                  "compress", f"{info['name']} ({ratio:.0f}% сжато)")
            except Exception as exc:
                stop_event.set()
                await progress_task
                logger.exception("Compress failed")
                await query.edit_message_text(
                    f"❌ Ошибка сжатия:\n`{exc}`", parse_mode="Markdown"
                )

        elif action == "back":
            ext = info["ext"]
            await query.edit_message_text(
                f"{fmt_emoji(ext)} *{info['name']}*\n\nЧто хочешь сделать?",
                reply_markup=action_keyboard(key, ext),
                parse_mode="Markdown",
            )
        return

    # ── cmp|level|key — выбор уровня сжатия ──
    if data.startswith("cmp|"):
        _, level, key = data.split("|", 2)
        info = context.bot_data.get(key)
        if not info:
            await query.edit_message_text("❌ Сессия устарела. Отправь файл заново.")
            return
        quality_map = {"light": 85, "medium": 60, "heavy": 30}
        context.user_data["compress_quality"] = quality_map.get(level, 60)
        # Эмулируем action=compress_do
        query.data = f"act|compress_do|{key}"
        await handle_callback(update, context)
        return

    # ── conv|key|tgt_ext ──
    if data.startswith("conv|"):
        _, key, tgt_ext = data.split("|", 2)
        info = context.bot_data.get(key)
        if not info:
            await query.edit_message_text("❌ Сессия устарела. Отправь файл заново.")
            return

        src_ext       = info["ext"]
        original_name = info["name"]

        # Запускаем прогресс-бар
        stop_event    = asyncio.Event()
        progress_task = asyncio.create_task(
            animated_progress(
                context, query.message.chat_id, query.message.message_id,
                f"Конвертирую в {tgt_ext.upper()}", stop_event,
            )
        )

        try:
            with tempfile.TemporaryDirectory() as tmpdir:
                tmp      = Path(tmpdir)
                src_path = tmp / f"input.{src_ext}"
                dst_path = tmp / f"output.{tgt_ext}"

                tg_file = await context.bot.get_file(info["file_id"])
                await tg_file.download_to_drive(str(src_path))

                result_path = await asyncio.get_event_loop().run_in_executor(
                    None, do_convert, src_path, src_ext, tgt_ext, dst_path
                )

                stop_event.set()
                await progress_task

                if not result_path.exists() or result_path.stat().st_size == 0:
                    await query.edit_message_text("❌ Конвертация завершилась пустым файлом.")
                    return

                stem     = Path(original_name).stem
                out_ext  = result_path.suffix.lstrip(".")
                out_name = f"{stem}.{out_ext}"

                await query.edit_message_text(
                    f"✅ Готово! Отправляю *{out_name}*…", parse_mode="Markdown"
                )
                with open(result_path, "rb") as f:
                    await context.bot.send_document(
                        chat_id=query.message.chat_id,
                        document=f,
                        filename=out_name,
                        caption=f"🎉 {original_name} → {out_name}",
                    )
                await query.delete_message()

                record_action(user.id, user.username or "",
                              "convert", f"{original_name} → {out_name}")

        except Exception as exc:
            stop_event.set()
            await progress_task
            logger.exception("Conversion failed")
            await query.edit_message_text(
                f"❌ Ошибка конвертации:\n`{exc}`", parse_mode="Markdown"
            )


# ── Rename conversation ────────────────────────────────────────────────────────

async def handle_new_name(update: Update, context: ContextTypes.DEFAULT_TYPE) -> int:
    new_name_input = update.message.text.strip()
    key  = context.user_data.get("rename_key")
    info = context.bot_data.get(key) if key else None
    user = update.effective_user

    if not info:
        await update.message.reply_text("❌ Сессия устарела. Отправь файл заново.")
        return ConversationHandler.END

    original_name = info["name"]
    original_ext  = Path(original_name).suffix

    new_name = (new_name_input + original_ext
                if "." not in Path(new_name_input).name[1:]
                else new_name_input)

    await update.message.reply_text(
        f"⏳ Переименовываю в *{new_name}*…", parse_mode="Markdown"
    )

    with tempfile.TemporaryDirectory() as tmpdir:
        src_path = Path(tmpdir) / original_name
        tg_file  = await context.bot.get_file(info["file_id"])
        await tg_file.download_to_drive(str(src_path))

        with open(src_path, "rb") as f:
            await update.message.reply_document(
                document=f,
                filename=new_name,
                caption=f"✏️ {original_name} → {new_name}",
            )

    record_action(user.id, user.username or "", "rename",
                  f"{original_name} → {new_name}")
    return ConversationHandler.END


# ── Main ───────────────────────────────────────────────────────────────────────

async def cleanup_sessions(context: ContextTypes.DEFAULT_TYPE) -> None:
    """Удаляет устаревшие сессии из bot_data (запускается каждые 30 мин)."""
    now     = time.time()
    expired = [k for k, v in context.bot_data.items()
               if isinstance(v, dict) and now - v.get("created_at", now) > SESSION_TTL]
    for k in expired:
        del context.bot_data[k]
    if expired:
        logger.info(f"🧹 Очищено сессий: {len(expired)}")


async def post_init(app: Application) -> None:
    # Запускаем автоочистку каждые 30 минут
    if app.job_queue:
        app.job_queue.run_repeating(cleanup_sessions, interval=1800, first=1800)
    await app.bot.set_my_commands([
        BotCommand("start",   "Начать / помощь"),
        BotCommand("help",    "Список конвертаций"),
        BotCommand("stats",   "Моя статистика"),
        BotCommand("history", "История операций"),
        BotCommand("cancel",  "Отменить действие"),
    ])


def main() -> None:
    token = os.environ.get("BOT_TOKEN")
    if not token:
        raise SystemExit("❌  Укажи BOT_TOKEN в .env файле.")

    app = Application.builder().token(token).post_init(post_init).build()

    rename_conv = ConversationHandler(
        entry_points=[CallbackQueryHandler(handle_callback, pattern=r"^act\|rename\|")],
        states={
            WAITING_NEW_NAME: [
                MessageHandler(filters.TEXT & ~filters.COMMAND, handle_new_name)
            ],
        },
        fallbacks=[CommandHandler("cancel", cmd_cancel)],
        per_message=False,
    )

    app.add_handler(CommandHandler("start",   cmd_start))
    app.add_handler(CommandHandler("help",    cmd_help))
    app.add_handler(CommandHandler("stats",   cmd_stats))
    app.add_handler(CommandHandler("history", cmd_history))
    app.add_handler(CommandHandler("admin",   cmd_admin))
    app.add_handler(CommandHandler("cancel",  cmd_cancel))
    app.add_handler(rename_conv)
    app.add_handler(MessageHandler(filters.Document.ALL, handle_document))
    app.add_handler(MessageHandler(filters.PHOTO, handle_photo))
    app.add_handler(CallbackQueryHandler(handle_callback))

    logger.info("🤖 Bot is running…")
    app.run_polling(allowed_updates=Update.ALL_TYPES)


if __name__ == "__main__":
    main()
